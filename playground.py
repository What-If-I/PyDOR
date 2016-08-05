import openpyxl
from dor_pars import *
from openpyxl.cell import get_column_letter, column_index_from_string


def get_header_fields(sheet):
    for cur_col in range(1, sheet.max_column + 1):
        cur_row = 1
        if sheet.cell(row=cur_row, column=cur_col).value is None:
            cur_row = 2
        yield sheet.cell(row=cur_row, column=cur_col).value.replace("\n", "").replace("\r", ""), \
            sheet.cell(row=cur_row, column=cur_col).column


errors = 0
directory = '//10.68.25.4/Project/UCP/Reports/Daily reports/'
os.chdir(directory)

# Список всех файлов в директории
reports = [(name, path) for name, path in reports_name_and_path(exclude_folder="_old")]

# Сегодняшний день и месяц
today = datetime.date.today()
current_month = "{:%B}".format(today)
yesterday = today - datetime.timedelta(1)
yesterday_day = yesterday.day

bsc_status_reports = find_report(reports, "BSC-Status_", today)

bsc_status_report_wb = openpyxl.load_workbook(bsc_status_reports)
bsc_status_report_sheet = bsc_status_report_wb.active

bsc_status_total_row = search_in_row(bsc_status_report_sheet, "Total:", 1,
                                     start=1, end=bsc_status_report_sheet.max_row).row

bsc_status_total_cell_coordinates = {"on_call": "E{}".format(bsc_status_total_row),
                                     "after_call": "G{}".format(bsc_status_total_row),
                                     "mail_flex": "H{}".format(bsc_status_total_row),
                                     "back_office_work": "J{}".format(bsc_status_total_row),
                                     "available": "D{}".format(bsc_status_total_row)
                                     }

bsc_status_statistic_col_coord = {value: coordinates for value, coordinates in get_header_fields(bsc_status_report_sheet)}

print(bsc_status_statistic_col_coord)
