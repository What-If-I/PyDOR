import os
import datetime
from openpyxl.cell import get_column_letter, column_index_from_string


def search_in_column(sheet, value, in_row, start, end):
    """
    Находит ближайшую ячейки по указанному содержимому.
    :param sheet: Название листа
    :param value: Искомое значение
    :param in_row: Поиск в строке
    :param start: Колонка начало поиска
    :param end: Колонка окончания поиска
    :return Координаты ячейки:
    """
    for column in range(start, end + 1):
        active_column = sheet.cell(row=in_row, column=column)
        if active_column.value is None:
            continue
        if active_column.value == value:
            return active_column
    return False


def search_in_row(sheet, value, in_column, start, end):
    """
    Находит ближайшую ячейки по указанному содержимому.
    :param sheet: Название листа
    :param value: Искомое значение
    :param in_column: Поиск в колонке
    :param start: Колонка начало поиска
    :param end: Колонка окончания поиска
    :return Координаты ячейки:
    """
    for row in range(start, end + 1):
        active_column = sheet.cell(row=row, column=in_column)
        if active_column.value == value:
            return active_column
    return False


def reports_name_and_path(exclude_folder=''):
    for dir_path, dir_names, file_names in os.walk(os.getcwd()):
        if exclude_folder in dir_names:
            dir_names.remove(exclude_folder)
        for file_name in file_names:
            yield file_name, dir_path


def find_report(reports, beginning, day):
    for report in reports:
        if report[0].startswith(beginning) and get_date(report[0]) == day:
            report_path = os.path.join(report[1], report[0])
            return report_path
    return None


def get_dor_sheet_and_day_column(dor, sheet, date, next_month_after=31):
    day = date.day
    if day <= next_month_after:
        month = "{:%B}".format(date)
    else:  # go to next month
        next_month = date + datetime.timedelta(15)
        month = "{:%B}".format(next_month)
    dor_sheet = dor.get_sheet_by_name(sheet)
    month_cell = search_in_column(dor_sheet, month, 1, start=1, end=dor_sheet.max_column)
    month_column_index = column_index_from_string(month_cell.column)
    day_cell = search_in_column(dor_sheet, day, 2, start=month_column_index, end=dor_sheet.max_column)
    column_index = column_index_from_string(day_cell.column)
    return dor_sheet, column_index


def get_sec(s):
    if None:
        return None
    times = s.split(':')
    seconds = 0
    i = 0
    for time in reversed(times):
        seconds += int(time) * 60 ** i
        i += 1
    return seconds


# TODO: не учитывает праздники. Надо поправить.
def is_weekend(date):
    if date.isoweekday() in range(6, 8):
        return True
    else:
        return False


def get_date(f):
    f_name = f.split('.')[0]
    f_date = datetime.datetime.strptime(f_name.split('_')[-1], "%Y-%m-%d-%H-%M-%S").date()
    return f_date


def get_header_fields(sheet):
    for cur_col in range(1, sheet.max_column + 1):
        cur_row = 1
        # go to next row right below Not Ready to not skip After Call auto
        if sheet.cell(row=cur_row, column=cur_col + 1).value is None:
            cur_row = 2
        cell_value = sheet.cell(row=cur_row, column=cur_col).value.replace("\n", " ").replace("\r", " ")
        cell_column = sheet.cell(row=cur_row, column=cur_col).column
        yield cell_value, cell_column


def get_status_total(sheet):
    name_col_number = {value: coordinates for value, coordinates in get_header_fields(sheet)}
    name_value = {name: sheet[coord+str(sheet.max_row)].value for name, coord in name_col_number.items()}
    return name_value
