import openpyxl
import logging
import sys
import getopt
import logger
from dor_pars import *


# noinspection SpellCheckingInspection,PyBroadException
def main():
    errors = 0
    directory = '//10.68.25.4/Project/UCP/Reports/Daily reports/'
    # directory = './Daily reports'
    os.chdir(directory)

    logging.basicConfig(filename='../script/logs/fill_dor_detailed-log.log', level=logging.DEBUG)
    error_logger = logger.Logger("../script/fill_dor_log.log", )

    opts, args = getopt.getopt(sys.argv[1:], "-d")  # принимаем день из консоли

    # Сегодняшняя дата
    today = datetime.date.today()

    # Собираем за указанный день
    for opt in opts:
        if '-d' in opt:
            today = datetime.datetime.strptime(args[0], "%d-%m-%y").date() + datetime.timedelta(1)

    yesterday = today - datetime.timedelta(1)
    current_month = "{:%B}".format(yesterday)
    yesterday_day = yesterday.day

    # отркываем файл DOR
    try:
        dor = openpyxl.load_workbook("DOR.xlsx")
    except FileNotFoundError:
        error_logger.append_error("Файл DOR не найден. Убедитесь что файл DOR.xlsx находится в директории")
        exit()

    # Список всех файлов в директории
    if today == datetime.date.today():
        reports = [(name, path) for name, path in reports_name_and_path("_old")]
    else:
        reports = [(name, path) for name, path in reports_name_and_path()]

    # ================================= Начало АА =================================================
    try:
        if not is_weekend(yesterday):
            # открываем страницу AA, находим столбец текущего дня
            dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "AA", yesterday)

            # находим файл с отчётами по АА
            aa_reports = find_report(reports, "AA_", today)

            # открываем отчёт AA
            aa_report = openpyxl.load_workbook(aa_reports)
            aa_report_sheet = aa_report.active
            aa_statistic = {"entered": aa_report_sheet['D9'].value,
                            "answered": aa_report_sheet['E9'].value,
                            "answered<sl": aa_report_sheet['G9'].value,
                            "abandoned": aa_report_sheet['F9'].value,
                            "AHT": get_sec(aa_report_sheet["P9"].value)
                            }

            # Заполняем DOR AA
            dor_sheet.cell(column=cur_day_column_index, row=5).value = aa_statistic["AHT"]
            dor_sheet.cell(column=cur_day_column_index, row=7).value = aa_statistic["entered"]
            dor_sheet.cell(column=cur_day_column_index, row=8).value = aa_statistic["answered"]
            dor_sheet.cell(column=cur_day_column_index, row=9).value = aa_statistic["answered<sl"]
            dor_sheet.cell(column=cur_day_column_index, row=10).value = aa_statistic["abandoned"]
    except:
        print("Не удалось заполнить АА")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ============================================================================================

    # ================ Начало АА Сервис =============================
    try:
        # открываем страницу AA, находим столбец текущего дня
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "AA", yesterday)

        # находим файл с отчётами по АА сервис
        aa_sc_reports = find_report(reports, "AA-Service-Centre_", today)

        # открываем отчёт AA-SC
        aa_sc = openpyxl.load_workbook(aa_sc_reports)
        aa_sc_sheet = aa_sc.active
        aa_sc_statistic = {"entered": aa_sc_sheet['D9'].value,
                           "answered": aa_sc_sheet['E9'].value,
                           "answered<sl": aa_sc_sheet['G9'].value,
                           "abandoned": aa_sc_sheet['F9'].value
                           }

        # Заполняем DOR AA Service Center
        dor_sheet.cell(column=cur_day_column_index, row=15).value = aa_sc_statistic["entered"]
        dor_sheet.cell(column=cur_day_column_index, row=16).value = aa_sc_statistic["answered"]
        dor_sheet.cell(column=cur_day_column_index, row=17).value = aa_sc_statistic["answered<sl"]
        dor_sheet.cell(column=cur_day_column_index, row=18).value = aa_sc_statistic["abandoned"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку АА")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ============================================================================================

    # ======================= Начало BSC =========================================================
    try:
        if not is_weekend(yesterday):
            # открываем страницу BSC, находим столбец текущего дня
            dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "BSC", yesterday)

            # находим файл с отчётами
            bsc_calls_reports = find_report(reports, "BSC_", today)
            bsc_status_reports = find_report(reports, "BSC-Status_", today)

            # открываем отчёт BSC
            bsc_report_wb = openpyxl.load_workbook(bsc_calls_reports)
            bsc_report_sheet = bsc_report_wb.active
            bsc_statistic = {"entered": bsc_report_sheet['D9'].value,
                             "answered": bsc_report_sheet['E9'].value,
                             "answered<sl": bsc_report_sheet['G9'].value,
                             "abandoned": bsc_report_sheet['F9'].value,
                             "ghost_calls": bsc_report_sheet['J9'].value,
                             "AHT": get_sec(bsc_report_sheet['P9'].value)
                             }

            # открываем отчёт Bosch-status и собираем статистику
            bsc_status_report_wb = openpyxl.load_workbook(bsc_status_reports)
            bsc_status_report_sheet = bsc_status_report_wb.active

            status_time = get_columns_total(bsc_status_report_sheet)

            # Заполняем DOR BSC

            dor_sheet.cell(column=cur_day_column_index, row=5).value = bsc_statistic["AHT"]
            dor_sheet.cell(column=cur_day_column_index, row=7).value = bsc_statistic["entered"]
            dor_sheet.cell(column=cur_day_column_index, row=8).value = bsc_statistic["answered"]
            dor_sheet.cell(column=cur_day_column_index, row=9).value = bsc_statistic["answered<sl"]
            dor_sheet.cell(column=cur_day_column_index, row=10).value = bsc_statistic["abandoned"]
            dor_sheet.cell(column=cur_day_column_index, row=11).value = bsc_statistic["ghost_calls"]

            dor_sheet.cell(column=cur_day_column_index, row=12).value = calc_occupancy(
                status_time["Available"],
                status_time["On Call"],
                status_time["After Call Work (auto)"],
                status_time["Mail Flex"],
                status_time["Back Office Work"]
            )

    except:
        error_logger.append_error("Не удалось заполнить вкладку BSC")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ===================================================================================================

    # ============================ Buderus ===============================================================
    try:
        if not is_weekend(yesterday):
            # открываем страницу Buderus, находим столбец текущего дня
            dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "Buderus", yesterday)

            # находим файл с отчётами
            buderus_reports = find_report(reports, "Buderus_", today)

            # открываем отчёт Buderus
            buderus_report_wb = openpyxl.load_workbook(buderus_reports)
            buderus_report_sheet = buderus_report_wb.active
            buderus_statistic = {"entered": buderus_report_sheet['D9'].value,
                                 "answered": buderus_report_sheet['E9'].value,
                                 "answered<sl": buderus_report_sheet['G9'].value,
                                 "abandoned": buderus_report_sheet['F9'].value,
                                 "AHT": get_sec(buderus_report_sheet['P9'].value)
                                 }

            # Заполняем DOR Buderus

            dor_sheet.cell(column=cur_day_column_index, row=5).value = buderus_statistic["AHT"]
            dor_sheet.cell(column=cur_day_column_index, row=7).value = buderus_statistic["entered"]
            dor_sheet.cell(column=cur_day_column_index, row=8).value = buderus_statistic["answered"]
            dor_sheet.cell(column=cur_day_column_index, row=9).value = buderus_statistic["answered<sl"]
            dor_sheet.cell(column=cur_day_column_index, row=10).value = buderus_statistic["abandoned"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку Buderus")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Buderus-Sales ===============================================================
    try:
        if not is_weekend(yesterday):
            # открываем страницу Buderus-Sales, находим столбец текущего дня
            dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "Buderus Sale", yesterday)

            # находим файл с отчётами
            buderus_sales_reports = find_report(reports, "Buderus-Sale_", today)

            # открываем отчёт Buderus-Sales
            buderus_sales_report_wb = openpyxl.load_workbook(buderus_sales_reports)
            buderus_sales_report_sheet = buderus_sales_report_wb.active
            buderus_sales_statistic = {"entered": buderus_sales_report_sheet['D9'].value,
                                       "answered": buderus_sales_report_sheet['E9'].value,
                                       "answered<sl": buderus_sales_report_sheet['G9'].value,
                                       "abandoned": buderus_sales_report_sheet['F9'].value,
                                       "AHT": get_sec(buderus_sales_report_sheet['P9'].value)
                                       }

            # Заполняем DOR Buderus-Sales

            dor_sheet.cell(column=cur_day_column_index, row=5).value = buderus_sales_statistic["AHT"]
            dor_sheet.cell(column=cur_day_column_index, row=7).value = buderus_sales_statistic["entered"]
            dor_sheet.cell(column=cur_day_column_index, row=8).value = buderus_sales_statistic["answered"]
            dor_sheet.cell(column=cur_day_column_index, row=9).value = buderus_sales_statistic["answered<sl"]
            dor_sheet.cell(column=cur_day_column_index, row=10).value = buderus_sales_statistic["abandoned"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку Buderus Sales")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ================================ Начало K2W Сервис =============================================
    try:
        # открываем страницу K2W, находим столбец текущего дня
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "K2W", yesterday)

        # находим файл с отчётами по K2W
        k2w_inbound_reports = find_report(reports, "K2W-calls_", today)
        k2w_outbound_reports = find_report(reports, "K2W-outbound-calls", today)

        # открываем входящие звонки K2W
        k2w_outbound_calls = openpyxl.load_workbook(k2w_inbound_reports)
        k2w_inbound_calls_sheet = k2w_outbound_calls.active
        k2w_inbound_calls_statistic = {"entered": k2w_inbound_calls_sheet['C6'].value,
                                       "answered": k2w_inbound_calls_sheet['D6'].value,
                                       "answered<sl": k2w_inbound_calls_sheet['E6'].value,
                                       "abandoned": k2w_inbound_calls_sheet['H6'].value,
                                       "abandoned<5": k2w_inbound_calls_sheet['I6'].value,
                                       "abandoned>5": k2w_inbound_calls_sheet['J6'].value,
                                       "AHT": k2w_inbound_calls_sheet['M6'].value
                                       }

        # открываем исходящие звонки K2W
        k2w_outbound_calls = openpyxl.load_workbook(k2w_outbound_reports)
        k2w_outbound_calls_sheet = k2w_outbound_calls.active
        k2w_outbound_calls_statistic = {"outbound-calls": k2w_outbound_calls_sheet['D7'].value,
                                        "AHT": k2w_outbound_calls_sheet['E7'].value,
                                        }

        # Заполняем DOR K2W
        dor_sheet.cell(column=cur_day_column_index, row=5).value = get_sec(k2w_inbound_calls_statistic["AHT"])
        dor_sheet.cell(column=cur_day_column_index, row=7).value = k2w_inbound_calls_statistic["entered"]
        dor_sheet.cell(column=cur_day_column_index, row=8).value = k2w_inbound_calls_statistic["answered"]
        dor_sheet.cell(column=cur_day_column_index, row=9).value = k2w_inbound_calls_statistic["answered<sl"]
        dor_sheet.cell(column=cur_day_column_index, row=10).value = k2w_inbound_calls_statistic["abandoned"]
        dor_sheet.cell(column=cur_day_column_index, row=11).value = k2w_inbound_calls_statistic["abandoned<5"]
        dor_sheet.cell(column=cur_day_column_index, row=12).value = k2w_inbound_calls_statistic["abandoned>5"]

        dor_sheet.cell(column=cur_day_column_index, row=14).value = k2w_outbound_calls_statistic["outbound-calls"]
        dor_sheet.cell(column=cur_day_column_index, row=15).value = get_sec(k2w_outbound_calls_statistic["AHT"])

    except:
        error_logger.append_error("Не удалось заполнить вкладку K2W")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ============================================================================================

    # ============================ Michelin ===============================================================
    try:
        # открываем страницу Michelin, находим столбец текущего дня
        michelin_dor_sheet = dor.get_sheet_by_name("Michelin")
        curr_month_cell = search_in_column(michelin_dor_sheet, current_month, 1,
                                           start=1, end=michelin_dor_sheet.max_column)
        curr_month_column_index = column_index_from_string(curr_month_cell.column)
        curr_day_cell = search_in_column(michelin_dor_sheet, yesterday_day, 2,
                                         start=curr_month_column_index, end=michelin_dor_sheet.max_column)
        cur_day_column_index = column_index_from_string(curr_day_cell.column)

        # находим файл с отчётами
        michelin_calls_reports = find_report(reports, "Michelin-Calls_", today)
        michelin_status_reports = find_report(reports, "Michelin-Status_", today)
        michelin_votes_reports = find_report(reports, "Michelin-Votes_", today)

        # открываем отчёт Michelin-calls и собираем статистику
        michelin_calls_report_wb = openpyxl.load_workbook(michelin_calls_reports)
        michelin_calls_report_sheet = michelin_calls_report_wb.active
        michelin_calls_statistic = {"entered": michelin_calls_report_sheet['C6'].value,
                                    "answered": michelin_calls_report_sheet['D6'].value,
                                    "answered<sl": michelin_calls_report_sheet['E6'].value,
                                    "abandoned": michelin_calls_report_sheet['H6'].value,
                                    "abandoned<5": michelin_calls_report_sheet['I6'].value,
                                    "AHT": get_sec(michelin_calls_report_sheet['M6'].value),
                                    }

        michelin_status_report_wb = openpyxl.load_workbook(michelin_status_reports)
        michelin_status_report_sheet = michelin_status_report_wb.active

        status_time = get_columns_total(michelin_status_report_sheet)

        # открываем отчёт Michelin-votes и собираем статистику
        michelin_votes_report_wb = openpyxl.load_workbook(michelin_votes_reports)
        michelin_votes_report_sheet = michelin_votes_report_wb.get_sheet_by_name("Kazan Michelin Voting Total")
        michelin_votes_statistic = {"vote1": michelin_votes_report_sheet['B5'].value,
                                    "vote2": michelin_votes_report_sheet['C5'].value,
                                    "vote3": michelin_votes_report_sheet['D5'].value,
                                    "vote4": michelin_votes_report_sheet['E5'].value,
                                    "vote5": michelin_votes_report_sheet['F5'].value
                                    }

        # Заполняем DOR Michelin

        michelin_dor_sheet.cell(column=cur_day_column_index, row=5).value = michelin_calls_statistic["AHT"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=7).value = michelin_calls_statistic["entered"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=8).value = michelin_calls_statistic["answered"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=9).value = michelin_calls_statistic["answered<sl"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=10).value = michelin_calls_statistic["abandoned"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=11).value = michelin_calls_statistic["abandoned<5"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=12).value = michelin_votes_statistic["vote1"] + \
                                                                             michelin_votes_statistic["vote2"]
        michelin_dor_sheet.cell(column=cur_day_column_index, row=13).value = sum(michelin_votes_statistic.values())

        michelin_dor_sheet.cell(column=cur_day_column_index, row=15).value = calc_occupancy(
            status_time["Available"],
            status_time["On Call"],
            status_time["After Call Work (auto)"],
            status_time["After Call Work (status)"],
            status_time["Mail Flex"],
            status_time["Back Office Work"]
        )
    except:
        error_logger.append_error("Не удалось заполнить вкладку Michelin")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Invitro ===============================================================
    try:
        # открываем страницу Invitro, находим столбец текущего дня
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "Invitro", yesterday)

        # находим файл с отчётами
        invitro_calls_reports = find_report(reports, "Invitro_Calls(Cities+Expert)_", today)
        invitro_cities_calls_reports = find_report(reports, "Invitro_Calls(Cities)_", today)
        invitro_status_reports = find_report(reports, "Invitro-Status_", today)

        # открываем отчёт Invitro-calls и собираем статистику
        invitro_calls_report_wb = openpyxl.load_workbook(invitro_calls_reports)
        invitro_calls_report_sheet = invitro_calls_report_wb.active
        invitro_calls_statistic = {"entered": invitro_calls_report_sheet['D6'].value,
                                   "answered": invitro_calls_report_sheet['E6'].value,
                                   "answered<sl": invitro_calls_report_sheet['F6'].value,
                                   "abandoned": invitro_calls_report_sheet['G6'].value,
                                   "abandoned<5": invitro_calls_report_sheet['H6'].value,
                                   "AHT": get_sec(invitro_calls_report_sheet['M6'].value),
                                   "ATT": get_sec(invitro_calls_report_sheet['L6'].value),
                                   "ACW": get_sec(invitro_calls_report_sheet['Q6'].value),
                                   }

        # открываем отчёт Invitro-cities-calls и собираем статистику
        invitro_cities_calls_reports_wb = openpyxl.load_workbook(invitro_cities_calls_reports)
        invitro_cities_calls_report_sheet = invitro_cities_calls_reports_wb.active
        invitro_cities_calls_statistic = {
            "AHT": get_sec(invitro_cities_calls_report_sheet['M6'].value),
            "ATT": get_sec(invitro_cities_calls_report_sheet['L6'].value),
            "ACW": get_sec(invitro_cities_calls_report_sheet['Q6'].value)
        }

        # открываем отчёт Invitro-status и собираем статистику
        invitro_status_report_wb = openpyxl.load_workbook(invitro_status_reports)
        invitro_status_report_sheet = invitro_status_report_wb.active

        status_time = get_columns_total(invitro_status_report_sheet)

        # Заполняем DOR Invitro

        dor_sheet.cell(column=cur_day_column_index, row=5).value = invitro_cities_calls_statistic["AHT"]
        dor_sheet.cell(column=cur_day_column_index, row=6).value = invitro_cities_calls_statistic["ATT"]
        dor_sheet.cell(column=cur_day_column_index, row=7).value = invitro_cities_calls_statistic["ACW"]

        dor_sheet.cell(column=cur_day_column_index, row=10).value = invitro_calls_statistic["entered"]
        dor_sheet.cell(column=cur_day_column_index, row=11).value = invitro_calls_statistic["answered"]
        dor_sheet.cell(column=cur_day_column_index, row=12).value = invitro_calls_statistic["answered<sl"]
        dor_sheet.cell(column=cur_day_column_index, row=13).value = invitro_calls_statistic["abandoned"]
        dor_sheet.cell(column=cur_day_column_index, row=14).value = invitro_calls_statistic["abandoned<5"]

        dor_sheet.cell(column=cur_day_column_index, row=15).value = calc_occupancy(
            status_time["Available"],
            status_time["On Call"],
            status_time["After Call Work (auto)"],
            status_time["After Call Work (status)"],
            status_time["Chat"],
            status_time["Back Office Work"]
        )

    except:
        error_logger.append_error("Не удалось заполнить вкладку Invitro")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Invitro-Expert ========================================================
    try:
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "Expert", yesterday)

        # находим файл с отчётами
        invitro_expert_calls_reports = find_report(reports, "Expert-Calls_", today)

        # открываем отчёт Invitro-Expert-calls и собираем статистику
        invitro_expert_calls_report_wb = openpyxl.load_workbook(invitro_expert_calls_reports)
        invitro_expert_calls_report_sheet = invitro_expert_calls_report_wb.active
        invitro_epxert_calls_statistic = {"entered": invitro_expert_calls_report_sheet['D6'].value,
                                          "answered": invitro_expert_calls_report_sheet['E6'].value,
                                          "answered<sl": invitro_expert_calls_report_sheet['F6'].value,
                                          "abandoned": invitro_expert_calls_report_sheet['G6'].value,
                                          "abandoned<5": invitro_expert_calls_report_sheet['H6'].value,
                                          "AHT": get_sec(invitro_expert_calls_report_sheet['M6'].value),
                                          "ATT": get_sec(invitro_expert_calls_report_sheet['L6'].value),
                                          "ACW": get_sec(invitro_expert_calls_report_sheet['Q6'].value),
                                          }

        # Заполняем DOR Invitro-Expert

        dor_sheet.cell(column=cur_day_column_index, row=5).value = invitro_epxert_calls_statistic["AHT"]
        dor_sheet.cell(column=cur_day_column_index, row=6).value = invitro_epxert_calls_statistic["ATT"]
        dor_sheet.cell(column=cur_day_column_index, row=7).value = invitro_epxert_calls_statistic["ACW"]

        dor_sheet.cell(column=cur_day_column_index, row=10).value = invitro_epxert_calls_statistic["entered"]
        dor_sheet.cell(column=cur_day_column_index, row=11).value = invitro_epxert_calls_statistic[
            "answered"]
        dor_sheet.cell(column=cur_day_column_index, row=12).value = invitro_epxert_calls_statistic[
            "answered<sl"]
        dor_sheet.cell(column=cur_day_column_index, row=13).value = invitro_epxert_calls_statistic[
            "abandoned"]
        dor_sheet.cell(column=cur_day_column_index, row=14).value = invitro_epxert_calls_statistic[
            "abandoned<5"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку Invitro Expert")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Kaspersky-B2C ===============================================================

    try:
        # открываем страницу Kaspersky-B2C, находим столбец текущего дня
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "B2C", yesterday, next_month_after=25)

        # находим файл с отчётами
        kaspersky_b2c_calls_reports = find_report(reports, "Kaspersky-B2C_", today)
        kaspersky_b2c_combined_status_reports = find_report(reports, "Kaspersky-B2C-Status(combined)_", today)
        kaspersky_b2c_agents_status_reports = find_report(reports, "Kaspersky-B2C-Status(agents)_", today)

        # открываем отчёт Kaspersky-B2C-calls и собираем статистику
        kaspersky_b2c_calls_report_wb = openpyxl.load_workbook(kaspersky_b2c_calls_reports)
        kaspersky_b2c_calls_report_sheet = kaspersky_b2c_calls_report_wb.active
        kaspersky_b2c_calls_statistic = {"AHT": get_sec(kaspersky_b2c_calls_report_sheet['P9'].value),
                                         "ATT": get_sec(kaspersky_b2c_calls_report_sheet['Q9'].value),
                                         "ACW": get_sec(kaspersky_b2c_calls_report_sheet['S9'].value),
                                         }

        # открываем отчёт Kaspersky-B2C-combined-status и собираем статистику
        kaspersky_b2c_combined_status_report_wb = openpyxl.load_workbook(kaspersky_b2c_combined_status_reports)
        kaspersky_b2c_combined_status_report_sheet = kaspersky_b2c_combined_status_report_wb.active

        status_time_combined = get_columns_total(kaspersky_b2c_combined_status_report_sheet)

        # открываем отчёт Kaspersky-B2C-combined-status и собираем статистику
        kaspersky_b2c_agents_status_report_wb = openpyxl.load_workbook(kaspersky_b2c_agents_status_reports)
        kaspersky_b2c_agents_status_report_sheet = kaspersky_b2c_agents_status_report_wb.active

        status_time_b2c = get_columns_total(kaspersky_b2c_agents_status_report_sheet)

        # Заполняем DOR Kaspersky-B2C

        dor_sheet.cell(column=cur_day_column_index, row=5).value = kaspersky_b2c_calls_statistic["AHT"]
        dor_sheet.cell(column=cur_day_column_index, row=6).value = kaspersky_b2c_calls_statistic["ATT"]
        dor_sheet.cell(column=cur_day_column_index, row=7).value = kaspersky_b2c_calls_statistic["ACW"]

        dor_sheet.cell(column=cur_day_column_index, row=14).value = calc_occupancy(
            status_time_combined["Available"],
            status_time_combined["On Call"],
            status_time_combined["After Call Work (auto)"],
            status_time_combined["After Call Work (status)"],
            status_time_combined["Admin Work"],
            status_time_combined["Available no ACD"]
        )

        dor_sheet.cell(column=cur_day_column_index, row=18).value = calc_occupancy(
            status_time_b2c["Available"],
            status_time_b2c["On Call"],
            status_time_b2c["After Call Work (auto)"],
            status_time_b2c["After Call Work (status)"],
            status_time_b2c["Admin Work"],
            status_time_b2c["Available no ACD"]
        )

        """ Occupancy
        (On call + after call + after call ручной + admin work + no ACD) /
        (On call + after call + after call ручной + admin work + no ACD + available)
        """
    except:
        error_logger.append_error("Не удалось заполнить вкладку Kaspersky B2C")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Kaspersky-B2B =====================================================

    try:
        if not is_weekend(yesterday):
            # открываем страницу Kaspersky-B2B, находим столбец текущего дня
            dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "B2B", yesterday, next_month_after=25)

            # находим файл с отчётами
            kaspersky_b2b_calls_reports = find_report(reports, "Kaspersky-B2B_", today)

            # открываем отчёт Kaspersky-B2B-calls и собираем статистику
            kaspersky_b2b_calls_report_wb = openpyxl.load_workbook(kaspersky_b2b_calls_reports)
            kaspersky_b2b_calls_report_sheet = kaspersky_b2b_calls_report_wb.active
            kaspersky_b2b_calls_statistic = {"AHT": get_sec(kaspersky_b2b_calls_report_sheet['P9'].value),
                                             "ATT": get_sec(kaspersky_b2b_calls_report_sheet['Q9'].value),
                                             "ACW": get_sec(kaspersky_b2b_calls_report_sheet['S9'].value),
                                             }

            # Заполняем DOR Kaspersky-B2B

            dor_sheet.cell(column=cur_day_column_index, row=5).value = kaspersky_b2b_calls_statistic["AHT"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку Kaspersky B2B")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    # ============================ Kaspersky-MEA ===============================================================

    try:
        # открываем страницу Kaspersky-MEA, находим столбец текущего дня
        dor_sheet, cur_day_column_index = get_dor_sheet_and_day_column(dor, "MEA", yesterday, next_month_after=25)

        # находим файл с отчётами
        kaspersky_mea_calls_reports = find_report(reports, "Kaspersky-MEA_", today)

        # открываем отчёт Kaspersky-MEA-calls и собираем статистику
        kaspersky_mea_calls_report_wb = openpyxl.load_workbook(kaspersky_mea_calls_reports)
        kaspersky_mea_calls_report_sheet = kaspersky_mea_calls_report_wb.active
        kaspersky_mea_calls_statistic = {
            "AHT": get_sec(kaspersky_mea_calls_report_sheet['P9'].value),
            "ATT": get_sec(kaspersky_mea_calls_report_sheet['Q9'].value),
            "ACW": get_sec(kaspersky_mea_calls_report_sheet['S9'].value),
        }

        # Заполняем DOR Kaspersky-MEA

        dor_sheet.cell(column=cur_day_column_index, row=6).value = kaspersky_mea_calls_statistic["AHT"]
    except:
        error_logger.append_error("Не удалось заполнить вкладку Kaspersky MEA")
        errors += 1
        logging.debug("\n==========================================")
        logging.exception("\nError occurred on %s \n" % datetime.datetime.today())

    # ================================================================================================

    try:
        dor.save("DOR.xlsx")
        dor.save("DOR_backup.xlsx")
    except PermissionError:
        errors += 1
        f_name = "DOR_copy_{:%d-%m-%y}.xlsx".format(today)
        logging.debug("\n==========================================")
        logging.exception("\nPermission denied. Saved as %s \n" % f_name)
        dor.save(f_name)
        error_logger.append_error("Не удалось перезаписать файл DOR, отчёт был сохранён как %s" % f_name)
    finally:
        error_logger.save()


if __name__ == '__main__':
    main()
